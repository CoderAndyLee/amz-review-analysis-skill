#!/usr/bin/env python3
"""Apply JSONL annotations onto reviews_master.xlsx and grow taxonomy columns."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from merge_reviews import HEADER_FILL, HEADER_FONT, WRAP

SEP = "|"
POS_FILL = PatternFill("solid", fgColor="E3F4E1")
NEG_FILL = PatternFill("solid", fgColor="FDECEC")
MIX_FILL = PatternFill("solid", fgColor="FFF4D6")


def col_key(polarity: str, l1: str, l2: str, l3: str = "") -> str:
    parts = [clean_label(polarity), clean_label(l1), clean_label(l2)]
    l3 = clean_label(l3)
    if l3:
        parts.append(l3)
    return SEP.join(parts)


def clean_label(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("|", "·")
    return text


def parse_col_key(key: str) -> dict:
    parts = [p.strip() for p in key.split(SEP)]
    while len(parts) < 4:
        parts.append("")
    polarity, l1, l2, l3 = parts[:4]
    leaf = 3 if l3 else 2 if l2 else 1
    return {
        "col_key": key,
        "polarity": polarity,
        "l1": l1,
        "l2": l2,
        "l3": l3,
        "leaf_level": leaf,
    }


def load_jsonl(path: Path) -> list[dict]:
    items = []
    text = path.read_text(encoding="utf-8")
    text = text.strip()
    if not text:
        return items
    if text[0] == "[":
        data = json.loads(text)
        if isinstance(data, list):
            return data
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        items.append(json.loads(line))
    return items


def find_header_map(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def ensure_columns(ws, keys: list[str]) -> dict[str, int]:
    header_map = find_header_map(ws)
    for key in keys:
        if key in header_map:
            continue
        col = ws.max_column + 1
        # if last header is empty, reuse
        if ws.cell(1, ws.max_column).value in (None, ""):
            col = ws.max_column
        cell = ws.cell(1, col, key)
        cell.fill = POS_FILL if key.startswith("好评") else NEG_FILL
        cell.font = Font(bold=True, name="PingFang SC", size=9, color="1F4E79")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 28
        header_map[key] = col
    return find_header_map(ws)


def append_excerpt(existing, excerpt: str) -> str:
    excerpt = (excerpt or "").strip()
    if not excerpt:
        return existing or ""
    existing = (existing or "").strip()
    if not existing:
        return excerpt
    if excerpt in existing:
        return existing
    return existing + " | " + excerpt


def apply(master_xlsx: Path, annotations_path: Path, batch_name: str) -> dict:
    wb = openpyxl.load_workbook(master_xlsx)
    ws = wb["reviews_master"]
    tax = wb["taxonomy"]

    annotations = load_jsonl(annotations_path)
    needed_keys = []
    parsed_anns = []
    for ann in annotations:
        assignments = ann.get("assignments") or ann.get("points") or []
        keys = []
        cleaned = []
        for a in assignments:
            polarity = clean_label(a.get("polarity") or a.get("极性") or "")
            l1 = clean_label(a.get("l1") or a.get("一级分类") or "")
            l2 = clean_label(a.get("l2") or a.get("二级分类") or "")
            l3 = clean_label(a.get("l3") or a.get("三级分类") or "")
            excerpt = (a.get("excerpt") or a.get("原文摘录") or "").strip()
            if polarity not in ("好评", "差评"):
                continue
            if not l1 or not excerpt:
                continue
            if not l2:
                # 指南：分析列至少落到二级，一级只作挂载
                l2 = "未细分"
            key = col_key(polarity, l1, l2, l3)
            keys.append(key)
            cleaned.append(
                {
                    "col_key": key,
                    "polarity": polarity,
                    "l1": l1,
                    "l2": l2,
                    "l3": l3,
                    "excerpt": excerpt,
                    "new": bool(a.get("new_column") or a.get("new")),
                }
            )
        needed_keys.extend(keys)
        parsed_anns.append((ann, cleaned))

    header_map = ensure_columns(ws, list(dict.fromkeys(needed_keys)))

    # row_id / review_key index
    row_by_id = {}
    row_by_key = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, header_map["row_id"]).value
        rkey = ws.cell(r, header_map["review_key"]).value
        if rid is not None:
            row_by_id[int(rid)] = r
        if rkey:
            row_by_key[str(rkey)] = r

    applied = 0
    missing = 0
    new_keys = set()
    for ann, cleaned in parsed_anns:
        r = None
        if ann.get("row_id") is not None:
            r = row_by_id.get(int(ann["row_id"]))
        if r is None and ann.get("review_key"):
            r = row_by_key.get(str(ann["review_key"]))
        if r is None:
            missing += 1
            continue
        mix = ann.get("sentiment_mix") or ann.get("情感") or ""
        if mix:
            ws.cell(r, header_map["sentiment_mix"]).value = mix
            fill = {"好评": POS_FILL, "差评": NEG_FILL, "混合": MIX_FILL}.get(mix)
            if fill:
                ws.cell(r, header_map["sentiment_mix"]).fill = fill
        ws.cell(r, header_map["annotate_status"]).value = "done"
        if batch_name:
            ws.cell(r, header_map["annotate_batch"]).value = batch_name
        for item in cleaned:
            c = header_map[item["col_key"]]
            cell = ws.cell(r, c)
            cell.value = append_excerpt(cell.value, item["excerpt"])
            cell.alignment = WRAP
            if item["new"]:
                new_keys.add(item["col_key"])
        applied += 1

    # rebuild taxonomy from current headers + counts
    cat_keys = [
        h
        for h in (ws.cell(1, c).value for c in range(1, ws.max_column + 1))
        if h and (str(h).startswith("好评|") or str(h).startswith("差评|"))
    ]
    existing_meta = {}
    for r in range(2, tax.max_row + 1):
        key = tax.cell(r, 1).value
        if not key:
            continue
        existing_meta[key] = {
            "status": tax.cell(r, 7).value or "active",
            "merged_into": tax.cell(r, 8).value or "",
            "first_row_id": tax.cell(r, 9).value,
            "first_asin": tax.cell(r, 10).value or "",
            "first_excerpt": tax.cell(r, 11).value or "",
        }

    # first occurrence + counts
    header_map = find_header_map(ws)
    stats = {}
    for key in cat_keys:
        col = header_map[key]
        first_row_id = None
        first_asin = ""
        first_excerpt = ""
        n_reviews = 0
        n_excerpts = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, col).value
            if not val:
                continue
            n_reviews += 1
            n_excerpts += len(str(val).split(" | "))
            if first_row_id is None:
                first_row_id = ws.cell(r, header_map["row_id"]).value
                first_asin = ws.cell(r, header_map["ASIN"]).value
                first_excerpt = str(val).split(" | ")[0]
        stats[key] = {
            "first_row_id": first_row_id,
            "first_asin": first_asin,
            "first_excerpt": first_excerpt,
            "n_reviews": n_reviews,
            "n_excerpts": n_excerpts,
        }

    if tax.max_row > 1:
        tax.delete_rows(2, tax.max_row - 1)
    for key in cat_keys:
        meta = parse_col_key(key)
        prev = existing_meta.get(key, {})
        st = stats.get(key, {})
        tax.append(
            [
                key,
                meta["polarity"],
                meta["l1"],
                meta["l2"],
                meta["l3"],
                meta["leaf_level"],
                prev.get("status", "active"),
                prev.get("merged_into", ""),
                st.get("first_row_id") or prev.get("first_row_id"),
                st.get("first_asin") or prev.get("first_asin"),
                (st.get("first_excerpt") or prev.get("first_excerpt") or "")[:300],
                st.get("n_reviews", 0),
                st.get("n_excerpts", 0),
            ]
        )
    for cell in tax[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    wb.save(master_xlsx)
    return {
        "applied": applied,
        "missing": missing,
        "new_columns": sorted(new_keys),
        "taxonomy_size": len(cat_keys),
        "batch": batch_name,
        "master": str(master_xlsx),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", required=True)
    parser.add_argument("--annotations", required=True)
    parser.add_argument("--batch", default="")
    args = parser.parse_args()
    result = apply(Path(args.master).expanduser(), Path(args.annotations).expanduser(), args.batch)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
