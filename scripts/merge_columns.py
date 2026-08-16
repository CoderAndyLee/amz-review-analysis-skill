#!/usr/bin/env python3
"""Merge synonym category columns. Excerpts move to the kept column."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from apply_annotations import append_excerpt, ensure_columns, find_header_map
from merge_reviews import HEADER_FILL, HEADER_FONT


def merge_pairs(master: Path, mapping: dict[str, str]):
    """mapping: old_col_key -> keep_col_key. Target columns are created if missing."""
    wb = openpyxl.load_workbook(master)
    ws = wb["reviews_master"]
    tax = wb["taxonomy"]
    header_map = ensure_columns(ws, list(mapping.values()))
    moved = 0
    skipped = []
    for old, keep in mapping.items():
        if old not in header_map:
            skipped.append(old)
            continue
        oc, kc = header_map[old], header_map[keep]
        if oc == kc:
            continue
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, oc).value
            if not val:
                continue
            cell = ws.cell(r, kc)
            cell.value = append_excerpt(cell.value, str(val))
            ws.cell(r, oc).value = None
            moved += 1

    # drop emptied source columns
    header_map = find_header_map(ws)
    drop_cols = []
    for old in mapping:
        col = header_map.get(old)
        if not col:
            continue
        if any(ws.cell(r, col).value for r in range(2, ws.max_row + 1)):
            continue
        drop_cols.append(col)
    for col in sorted(set(drop_cols), reverse=True):
        ws.delete_cols(col, 1)

    # rebuild taxonomy from remaining headers
    from apply_annotations import parse_col_key

    header_map = find_header_map(ws)
    cat_keys = [
        h
        for h in (ws.cell(1, c).value for c in range(1, ws.max_column + 1))
        if h and (str(h).startswith("好评|") or str(h).startswith("差评|"))
    ]
    if tax.max_row > 1:
        tax.delete_rows(2, tax.max_row - 1)
    for key in cat_keys:
        meta = parse_col_key(key)
        col = header_map[key]
        first_row_id = first_asin = first_excerpt = None
        n_reviews = n_excerpts = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, col).value
            if not val:
                continue
            n_reviews += 1
            n_excerpts += len(str(val).split(" | "))
            if first_row_id is None:
                first_row_id = ws.cell(r, header_map["row_id"]).value
                first_asin = ws.cell(r, header_map["ASIN"]).value
                first_excerpt = str(val).split(" | ")[0]
        tax.append(
            [
                key,
                meta["polarity"],
                meta["l1"],
                meta["l2"],
                meta["l3"],
                meta["leaf_level"],
                "active",
                "",
                first_row_id,
                first_asin,
                (first_excerpt or "")[:300],
                n_reviews,
                n_excerpts,
            ]
        )
    for cell in tax[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    wb.save(master)
    return {"moved_cells": moved, "pairs": mapping, "skipped_missing": skipped, "dropped_cols": len(set(drop_cols))}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", required=True)
    parser.add_argument("--map-json", required=True, help='{"旧列":"保留列", ...}')
    args = parser.parse_args()
    mapping = json.loads(Path(args.map_json).read_text(encoding="utf-8"))
    print(json.dumps(merge_pairs(Path(args.master), mapping), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
