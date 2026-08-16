#!/usr/bin/env python3
"""Merge SellerSprite review workbooks and join product fields."""

from __future__ import annotations

import argparse
import hashlib
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REVIEW_COLS = [
    "ASIN",
    "标题",
    "标题(翻译)",
    "内容",
    "内容(翻译)",
    "VP评论",
    "Vine Voice评论",
    "型号",
    "星级",
    "赞同数",
    "图片数量",
    "图片地址",
    "是否有视频",
    "视频地址",
    "评论链接",
    "评论人",
    "头像地址",
    "所属国家",
    "评论人主页",
    "红人计划链接",
    "评论时间",
]

PRODUCT_KEEP = {
    "ASIN": "ASIN",
    "品牌": "品牌",
    "商品标题": "商品标题",
    "父ASIN": "父ASIN",
    "价格($)": "价格",
    "小类BSR": "小类BSR",
    "月销量": "月销量",
    "月销售额($)": "月销售额",
    "评分": "站点评分",
    "评分数": "站点评分数",
    "#": "BSR位次",
    "上架时间": "上架时间",
}

ACCESSORY_MODEL_RE = re.compile(
    r"(blades?\s*set|set name:\s*0?4?garage\b|headlight|traction wheel|"
    r"wheel brushes|boundary wire|cutting disc)",
    re.I,
)
TRUE_TOKENS = {"Y", "YES", "TRUE", "1", "是"}


def load_sheet_rows(path: Path, sheet_hint: str | None = None) -> tuple[list, list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_hint and sheet_hint in wb.sheetnames:
        ws = wb[sheet_hint]
    else:
        ws = None
        for name in wb.sheetnames:
            if name.lower() != "note":
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = list(rows[0])
    return header, rows[1:]


def yn(value) -> str:
    if value is None:
        return ""
    return "Y" if str(value).strip().upper() in TRUE_TOKENS else ""


def review_key(asin, title, content, date, link) -> str:
    if link and str(link).strip().startswith("http"):
        return str(link).strip()
    raw = f"{asin}|{title}|{content}|{date}"
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()


def load_products(path: Path) -> dict[str, dict]:
    header, rows = load_sheet_rows(path, sheet_hint="US")
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for row in rows:
        asin = row[idx.get("ASIN", 0)]
        if not asin:
            continue
        rec = {}
        for src, dst in PRODUCT_KEEP.items():
            rec[dst] = row[idx[src]] if src in idx else None
        out[str(asin).strip()] = rec
    return out


def iter_review_files(reviews_dir: Path):
    files = sorted(reviews_dir.glob("*.xlsx"))
    files = [f for f in files if not f.name.startswith("~$")]
    return files


def merge(reviews_dir: Path, products_xlsx: Path | None) -> list[dict]:
    products = load_products(products_xlsx) if products_xlsx else {}
    merged = []
    seen = set()
    for f in iter_review_files(reviews_dir):
        header, rows = load_sheet_rows(f)
        idx = {h: i for i, h in enumerate(header)}

        def cell(row, name, default=""):
            i = idx.get(name)
            if i is None or i >= len(row):
                return default
            v = row[i]
            return default if v is None else v

        listing_asin = f.name.split("-")[0]
        for row in rows:
            review_asin = str(cell(row, "ASIN") or "").strip()
            asin = review_asin or listing_asin
            if not asin:
                continue
            title = cell(row, "标题")
            content = cell(row, "内容")
            date = cell(row, "评论时间")
            link = cell(row, "评论链接")
            key = review_key(asin, title, content, date, link)
            if key in seen:
                continue
            seen.add(key)
            prod = products.get(review_asin) or products.get(listing_asin) or {}
            model = str(cell(row, "型号") or "")
            country = str(cell(row, "所属国家") or "").strip()
            rec = {
                "row_id": None,
                "source_file": f.name,
                "listing_asin": listing_asin,
                "review_key": key,
                "ASIN": asin,
                "品牌": prod.get("品牌") or "",
                "商品标题": prod.get("商品标题") or "",
                "父ASIN": prod.get("父ASIN") or "",
                "BSR位次": prod.get("BSR位次"),
                "小类BSR": prod.get("小类BSR"),
                "价格": prod.get("价格"),
                "月销量": prod.get("月销量"),
                "月销售额": prod.get("月销售额"),
                "站点评分": prod.get("站点评分"),
                "站点评分数": prod.get("站点评分数"),
                "上架时间": prod.get("上架时间") or "",
                "标题": title or "",
                "标题(翻译)": cell(row, "标题(翻译)") or "",
                "内容": content or "",
                "内容(翻译)": cell(row, "内容(翻译)") or "",
                "型号": model,
                "星级": cell(row, "星级"),
                "赞同数": cell(row, "赞同数") or 0,
                "评论时间": date or "",
                "所属国家": country,
                "评论链接": link or "",
                "评论人": cell(row, "评论人") or "",
                "VP评论": yn(cell(row, "VP评论")),
                "Vine Voice评论": yn(cell(row, "Vine Voice评论")),
                "图片数量": cell(row, "图片数量") or 0,
                "图片地址": cell(row, "图片地址") or "",
                "是否有视频": yn(cell(row, "是否有视频")),
                "视频地址": cell(row, "视频地址") or "",
                "评论人主页": cell(row, "评论人主页") or "",
                "头像地址": cell(row, "头像地址") or "",
                "红人计划链接": cell(row, "红人计划链接") or "",
                "flag_non_us": "Y" if country and country.upper() != "US" else "",
                "flag_accessory_model": "Y" if ACCESSORY_MODEL_RE.search(model) else "",
                "flag_vp": yn(cell(row, "VP评论")),
                "flag_vine": yn(cell(row, "Vine Voice评论")),
                "flag_has_image": "Y" if (cell(row, "图片数量") or 0) not in ("", 0, "0") else "",
                "flag_has_video": yn(cell(row, "是否有视频")) or ("Y" if cell(row, "视频地址") else ""),
                "sentiment_mix": "",
                "annotate_status": "pending",
                "annotate_batch": "",
            }
            merged.append(rec)

    def sort_key(r):
        star = r["星级"] if isinstance(r["星级"], (int, float)) else 99
        return (str(r["品牌"]), str(r["ASIN"]), star, str(r["评论时间"]))

    merged.sort(key=sort_key)
    for i, rec in enumerate(merged, 1):
        rec["row_id"] = i
    return merged


MASTER_HEADERS = [
    "row_id",
    "source_file",
    "listing_asin",
    "review_key",
    "ASIN",
    "品牌",
    "商品标题",
    "父ASIN",
    "BSR位次",
    "小类BSR",
    "价格",
    "月销量",
    "月销售额",
    "站点评分",
    "站点评分数",
    "上架时间",
    "标题",
    "标题(翻译)",
    "内容",
    "内容(翻译)",
    "型号",
    "星级",
    "赞同数",
    "评论时间",
    "所属国家",
    "评论链接",
    "评论人",
    "VP评论",
    "Vine Voice评论",
    "图片数量",
    "图片地址",
    "是否有视频",
    "视频地址",
    "评论人主页",
    "头像地址",
    "红人计划链接",
    "flag_non_us",
    "flag_accessory_model",
    "flag_vp",
    "flag_vine",
    "flag_has_image",
    "flag_has_video",
    "sentiment_mix",
    "annotate_status",
    "annotate_batch",
]

TAXONOMY_HEADERS = [
    "col_key",
    "polarity",
    "l1",
    "l2",
    "l3",
    "leaf_level",
    "status",
    "merged_into",
    "first_row_id",
    "first_asin",
    "first_excerpt",
    "n_reviews",
    "n_excerpts",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="PingFang SC", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_FILL_ALT = PatternFill("solid", fgColor="F7F9FC")


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 8
        for cell in col[:80]:
            if cell.value is None:
                continue
            longest = max(longest, min(max_width, len(str(cell.value))))
        ws.column_dimensions[letter].width = longest + 2


def write_master(path: Path, records: list[dict], extra_headers: list[str] | None = None):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "reviews_master"
    headers = MASTER_HEADERS + (extra_headers or [])
    ws.append(headers)
    for rec in records:
        ws.append([rec.get(h, "") for h in headers])
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
        if row[0].row % 2 == 0:
            for cell in row:
                if not cell.fill or cell.fill.fgColor is None or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = THIN_FILL_ALT
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["P"].width = 36  # 标题
    ws.column_dimensions["R"].width = 56  # 内容
    ws.row_dimensions[1].height = 22

    tax = wb.create_sheet("taxonomy")
    tax.append(TAXONOMY_HEADERS)
    for cell in tax[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    tax.freeze_panes = "A2"
    autosize(tax)

    cov = wb.create_sheet("asin_coverage")
    cov.append(
        [
            "ASIN",
            "品牌",
            "商品标题",
            "BSR位次",
            "站点评分数",
            "已下载条数",
            "覆盖率",
            "有正文",
            "非美",
            "配件型号",
            "VP",
            "Vine",
            "1星",
            "2星",
            "3星",
            "4星",
            "5星",
        ]
    )
    by_asin: dict[str, list[dict]] = {}
    for rec in records:
        by_asin.setdefault(rec["ASIN"], []).append(rec)
    for asin, items in sorted(by_asin.items(), key=lambda kv: (kv[1][0].get("BSR位次") or 9999)):
        listed = items[0].get("站点评分数") or 0
        n = len(items)
        stars = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for it in items:
            try:
                s = int(it.get("星级") or 0)
            except (TypeError, ValueError):
                s = 0
            if s in stars:
                stars[s] += 1
        cov.append(
            [
                asin,
                items[0].get("品牌"),
                items[0].get("商品标题"),
                items[0].get("BSR位次"),
                listed,
                n,
                round(n / listed, 4) if listed else None,
                sum(1 for it in items if str(it.get("内容") or "").strip()),
                sum(1 for it in items if it.get("flag_non_us") == "Y"),
                sum(1 for it in items if it.get("flag_accessory_model") == "Y"),
                sum(1 for it in items if it.get("flag_vp") == "Y"),
                sum(1 for it in items if it.get("flag_vine") == "Y"),
                stars[1],
                stars[2],
                stars[3],
                stars[4],
                stars[5],
            ]
        )
    for cell in cov[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    cov.freeze_panes = "A2"
    autosize(cov, max_width=48)

    note = wb.create_sheet("readme")
    note.append(["field", "meaning"])
    notes = [
        ("reviews_master", "合并后的评论总表。源表字段在前，标注字段在后，分类列会在分析过程中追加。"),
        ("taxonomy", "活分类词典。一级/二级/三级在分析中长出，不预先锁死。"),
        ("asin_coverage", "每个 ASIN 的下载覆盖率与星级分布。"),
        ("col_key", "分类列名：极性|一级|二级 或 极性|一级|二级|三级"),
        ("sentiment_mix", "好评 / 差评 / 混合 / 信息不足"),
        ("flag_accessory_model", "型号字段像配件（刀片/车库等），主分析默认排除"),
        ("flag_non_us", "评论所属国家不是 US"),
        ("annotate_status", "pending / done"),
    ]
    for a, b in notes:
        note.append([a, b])
    for cell in note[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    autosize(note, max_width=80)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Merge SellerSprite review xlsx files.")
    parser.add_argument("--reviews-dir", required=True)
    parser.add_argument("--products-xlsx", default="")
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    reviews_dir = Path(args.reviews_dir).expanduser()
    products = Path(args.products_xlsx).expanduser() if args.products_xlsx else None
    out = Path(args.out).expanduser()
    records = merge(reviews_dir, products)
    write_master(out, records)
    print(f"merged={len(records)} out={out}")


if __name__ == "__main__":
    main()
