#!/usr/bin/env python3
"""Frequency tables from labeled reviews_master.xlsx."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


def stats(master: Path) -> dict:
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    ws = wb["reviews_master"]
    rows = list(ws.iter_rows(values_only=True))
    tax_rows = list(wb["taxonomy"].iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    idx = {h: i for i, h in enumerate(header)}
    cat_cols = [h for h in header if h and (str(h).startswith("好评|") or str(h).startswith("差评|"))]
    recs = []
    for row in rows[1:]:
        rec = {h: row[i] if i < len(row) else None for h, i in idx.items()}
        recs.append(rec)

    labeled = [r for r in recs if r.get("annotate_status") == "done"]
    main = [r for r in labeled if r.get("flag_accessory_model") != "Y"]

    def count_cols(subset, cols):
        out = []
        for col in cols:
            n = sum(1 for r in subset if r.get(col))
            if n:
                parts = col.split("|")
                out.append(
                    {
                        "col_key": col,
                        "polarity": parts[0] if parts else "",
                        "l1": parts[1] if len(parts) > 1 else "",
                        "l2": parts[2] if len(parts) > 2 else "",
                        "l3": parts[3] if len(parts) > 3 else "",
                        "n": n,
                    }
                )
        out.sort(key=lambda x: -x["n"])
        return out

    l1 = Counter()
    for r in main:
        for col in cat_cols:
            if r.get(col):
                parts = col.split("|")
                if len(parts) >= 2:
                    l1[f"{parts[0]}|{parts[1]}"] += 1

    by_brand = {}
    for brand in sorted({r.get("品牌") or "NA" for r in main}):
        subset = [r for r in main if (r.get("品牌") or "NA") == brand]
        by_brand[brand] = {
            "n": len(subset),
            "mix": dict(Counter(r.get("sentiment_mix") or "" for r in subset)),
            "top": count_cols(subset, cat_cols)[:12],
        }

    by_asin = {}
    for asin in sorted({r.get("listing_asin") or r.get("ASIN") for r in main}):
        subset = [r for r in main if (r.get("listing_asin") or r.get("ASIN")) == asin]
        if not subset:
            continue
        by_asin[asin] = {
            "n": len(subset),
            "brand": subset[0].get("品牌"),
            "title": subset[0].get("商品标题"),
            "mix": dict(Counter(r.get("sentiment_mix") or "" for r in subset)),
            "top": count_cols(subset, cat_cols)[:8],
        }

    return {
        "total": len(recs),
        "labeled": len(labeled),
        "main_labeled": len(main),
        "accessory_labeled": sum(1 for r in labeled if r.get("flag_accessory_model") == "Y"),
        "mix": dict(Counter(r.get("sentiment_mix") or "" for r in main)),
        "stars": dict(Counter(r.get("星级") for r in main)),
        "vp": sum(1 for r in main if r.get("flag_vp") == "Y"),
        "vine": sum(1 for r in main if r.get("flag_vine") == "Y"),
        "non_us": sum(1 for r in main if r.get("flag_non_us") == "Y"),
        "l1": [{"key": k, "n": v} for k, v in l1.most_common()],
        "leaves": count_cols(main, cat_cols),
        "by_brand": by_brand,
        "by_asin": by_asin,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    data = stats(Path(args.master))
    Path(args.out).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"labeled={data['labeled']} main={data['main_labeled']} leaves={len(data['leaves'])} out={args.out}")


if __name__ == "__main__":
    main()
