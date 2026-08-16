#!/usr/bin/env python3
"""Build delivery xlsx + board-data.js + copy the HTML shell."""

from __future__ import annotations

import argparse
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = PatternFill("solid", fgColor="1F4E79")
NAVY_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
TITLE_FONT = Font(bold=True, name="Calibri", size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")
POS = PatternFill("solid", fgColor="E3F4E1")
NEG = PatternFill("solid", fgColor="FDECEC")
LVL_FILL = {
    "极性": PatternFill("solid", fgColor="1F4E79"),
    "一级": PatternFill("solid", fgColor="2E75B6"),
    "二级": PatternFill("solid", fgColor="D6EAF8"),
    "三级": PatternFill("solid", fgColor="EEF6FB"),
    "摘录": PatternFill("solid", fgColor="FFFFFF"),
}
LVL_FONT = {
    "极性": Font(color="FFFFFF", bold=True, name="Calibri", size=11),
    "一级": Font(color="FFFFFF", bold=True, name="Calibri", size=10),
    "二级": Font(bold=True, name="Calibri", size=10, color="1F4E79"),
    "三级": Font(bold=True, name="Calibri", size=10, color="2E75B6"),
    "摘录": Font(name="Calibri", size=10),
}

SECTION_SPEC = [
    ("s01", "01", "买家最喜欢什么", "好评", "样本里最稳的正面，按一级频次降序。"),
    ("s02", "02", "差评主要抱怨什么", "差评", "差评一级按频次降序。"),
    ("s03", "03", "买家希望哪里更好", "差评", "从差评一级看改进空间。"),
    ("s04", "04", "哪些人群 / 场景更适合", "好评", "从好评一级归纳适用人群与场景。"),
    ("s05", "05", "哪些情况不适合", "差评", "从差评一级归纳不适用条件。"),
    ("s06", "06", "值不值这个价", "both", "性价比相关列优先，其余按频次补齐。"),
]


def header_row(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = NAVY
        cell.font = NAVY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 22


def autosize(ws, max_width=36, scan=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 10
        for cell in col[:scan]:
            if cell.value is None:
                continue
            longest = max(longest, min(max_width, len(str(cell.value))))
        ws.column_dimensions[letter].width = longest + 2


def sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return [], []
    rows = list(wb[name].iter_rows(values_only=True))
    if not rows:
        return [], []
    header = list(rows[0])
    recs = [{header[i]: row[i] if i < len(row) else None for i in range(len(header))} for row in rows[1:]]
    return header, recs


def load_master(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    header, recs = sheet_rows(wb, "reviews_master")
    if not header:
        name = wb.sheetnames[0]
        header, recs = sheet_rows(wb, name)
    tax_h, tax_recs = sheet_rows(wb, "taxonomy")
    cov_h, cov_recs = sheet_rows(wb, "asin_coverage")
    wb.close()
    return header, recs, tax_h, tax_recs, cov_h, cov_recs


def main_set(recs):
    return [r for r in recs if r.get("flag_accessory_model") != "Y"]


def cat_cols(header):
    return [h for h in header if h and (str(h).startswith("好评|") or str(h).startswith("差评|"))]


def parse_col(col):
    parts = str(col).split("|")
    while len(parts) < 4:
        parts.append("")
    return parts[0], parts[1], parts[2], parts[3]


def split_excerpts(val):
    return [p.strip() for p in str(val).split(" | ") if p.strip()]


def build_tree(header, recs):
    main = main_set(recs)
    cols = cat_cols(header)
    tree = {pol: {"row_ids": set(), "l1": {}} for pol in ("好评", "差评")}
    for rec in main:
        rid = rec.get("row_id")
        for col in cols:
            val = rec.get(col)
            if not val:
                continue
            pol, l1, l2, l3 = parse_col(col)
            if pol not in tree or not l1:
                continue
            pnode = tree[pol]
            pnode["row_ids"].add(rid)
            l1n = pnode["l1"].setdefault(l1, {"row_ids": set(), "l2": {}})
            l1n["row_ids"].add(rid)
            l2n = l1n["l2"].setdefault(l2 or "未细分", {"row_ids": set(), "l3": {}, "excerpts": []})
            l2n["row_ids"].add(rid)
            payload = {
                "row_id": rid,
                "brand": rec.get("品牌") or "",
                "star": rec.get("星级"),
                "asin": rec.get("listing_asin") or rec.get("ASIN") or "",
                "excerpts": split_excerpts(val),
            }
            if l3:
                l3n = l2n["l3"].setdefault(l3, {"row_ids": set(), "excerpts": []})
                l3n["row_ids"].add(rid)
                l3n["excerpts"].append(payload)
            else:
                l2n["excerpts"].append(payload)
    return tree


def sorted_items(mapping, key="row_ids"):
    return sorted(mapping.items(), key=lambda kv: (-len(kv[1][key]), kv[0]))


def ordered_cat_cols(header, recs):
    tree = build_tree(header, recs)
    ordered = []
    for pol in ("好评", "差评"):
        for l1, l1n in sorted_items(tree[pol]["l1"]):
            for l2, l2n in sorted_items(l1n["l2"]):
                if l2n["l3"]:
                    for l3, _l3n in sorted_items(l2n["l3"]):
                        ordered.append(f"{pol}|{l1}|{l2}|{l3}")
                if l2n["excerpts"] or not l2n["l3"]:
                    col = f"{pol}|{l1}|{l2}"
                    if col in header:
                        ordered.append(col)
    existing = set(cat_cols(header))
    ordered = [c for c in ordered if c in existing]
    leftover = [c for c in cat_cols(header) if c not in ordered]
    base = [h for h in header if h not in existing]
    return base + ordered + leftover, tree


def star_stats(recs):
    stars = Counter()
    for r in recs:
        try:
            s = int(r.get("星级") or 0)
        except (TypeError, ValueError):
            s = 0
        if s in (1, 2, 3, 4, 5):
            stars[s] += 1
    n = sum(stars.values()) or 1
    weighted = sum(k * stars[k] for k in stars) / n
    return stars, n, round(weighted, 1)


def write_readme(wb, meta, recs):
    main = main_set(recs)
    stars, n_star, rating = star_stats(main)
    mix = Counter(r.get("sentiment_mix") or "" for r in main)
    ws = wb.active
    ws.title = "00_说明"
    title = f"{meta['category']} Review 分析交付簿"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    rows = [
        ("样本", f"{meta.get('marketplace', '')} {meta.get('category', '')}；日期 {meta.get('date', '')}"),
        ("总行数", f"{len(recs)} 条（含配件旗标行）"),
        ("主分析口径", f"排除 flag_accessory_model=Y，主分析 {len(main)} 条"),
        ("综合评分", f"{rating} / 5"),
        ("星级", " · ".join(f"{s}★ {stars[s]}" for s in (5, 4, 3, 2, 1))),
        ("sentiment_mix", f"混合 {mix.get('混合', 0)} · 好评 {mix.get('好评', 0)} · 差评 {mix.get('差评', 0)} · 信息不足 {mix.get('信息不足', 0)}"),
        ("分类列", "极性|一级|二级 或 极性|一级|二级|三级；单元格为买家原文摘录"),
        ("不是普查", "抬头写清下载覆盖和未进池的头部 ASIN，禁止写成全类目结论"),
        ("看板", "4 Reports/review-analysis.html + board-data.js"),
    ]
    ws.append(["字段", "说明"])
    for a, b in rows:
        ws.append([a, b])
    for cell in ws[2]:
        cell.fill = NAVY
        cell.font = NAVY_FONT
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 88
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        row[1].alignment = WRAP


def write_master(wb, header, recs):
    ws = wb.create_sheet("01_reviews_master")
    header, _tree = ordered_cat_cols(header, recs)
    header_row(ws, header)
    for rec in recs:
        ws.append([rec.get(h) for h in header])
    for i, h in enumerate(header, 1):
        if not h:
            continue
        if str(h).startswith("好评|"):
            ws.cell(1, i).fill = POS
            ws.cell(1, i).font = Font(bold=True, color="1F4E79", size=9)
        elif str(h).startswith("差评|"):
            ws.cell(1, i).fill = NEG
            ws.cell(1, i).font = Font(bold=True, color="1F4E79", size=9)
        if h == "标题":
            ws.column_dimensions[get_column_letter(i)].width = 32
        elif h == "内容":
            ws.column_dimensions[get_column_letter(i)].width = 48
        elif str(h).startswith("好评|") or str(h).startswith("差评|"):
            ws.column_dimensions[get_column_letter(i)].width = 18
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    ws.freeze_panes = "E2"


def write_taxonomy(wb, tax_h, tax_recs):
    if not tax_h:
        return
    ws = wb.create_sheet("02_taxonomy")
    header_row(ws, tax_h)
    for rec in tax_recs:
        ws.append([rec.get(h) for h in tax_h])
    autosize(ws, max_width=42)


def write_l1(wb, header, recs):
    cols = cat_cols(header)
    main = main_set(recs)
    n_main = len(main) or 1
    l1 = Counter()
    for r in main:
        seen = set()
        for c in cols:
            if r.get(c):
                parts = str(c).split("|")
                if len(parts) >= 2:
                    seen.add(f"{parts[0]}|{parts[1]}")
        for k in seen:
            l1[k] += 1
    ws = wb.create_sheet("03_一级频次")
    header_row(ws, ["极性", "一级", "n", "占主分析比例"])
    for key, n in l1.most_common():
        pol, name = key.split("|", 1)
        ws.append([pol, name, n, round(n / n_main, 4)])
    autosize(ws)


def write_l2(wb, tax_recs):
    if not tax_recs:
        return
    ws = wb.create_sheet("04_二级频次")
    header_row(ws, ["col_key", "极性", "一级", "二级", "三级", "leaf_level", "n_reviews"])
    rows = sorted(tax_recs, key=lambda r: -(r.get("n_reviews") or 0))
    for r in rows:
        ws.append([r.get(k) for k in ["col_key", "polarity", "l1", "l2", "l3", "leaf_level", "n_reviews"]])
    autosize(ws, max_width=46)


def write_stars(wb, recs):
    main = main_set(recs)
    stars, n, weighted = star_stats(main)
    ws = wb.create_sheet("05_星级分布")
    header_row(ws, ["星级", "数量", "占比"])
    for s in (5, 4, 3, 2, 1):
        ws.append([s, stars[s], round(stars[s] / n, 4)])
    ws.append([])
    ws.append(["主分析条数", n, ""])
    ws.append(["加权均分", weighted, ""])
    autosize(ws, max_width=48)


def write_brand(wb, header, recs):
    cols = cat_cols(header)
    main = main_set(recs)
    tree = build_tree(header, recs)
    neg_l1 = [name for name, _ in sorted_items(tree["差评"]["l1"])]
    ws = wb.create_sheet("06_品牌切片")
    heads = ["品牌", "n", "好评", "差评", "混合", "信息不足"] + [f"差评|{n}" for n in neg_l1]
    header_row(ws, heads)
    brands = []
    seen = set()
    for r in main:
        b = r.get("品牌") or ""
        if b and b not in seen:
            seen.add(b)
            brands.append(b)
    brands.sort(key=lambda b: -sum(1 for r in main if r.get("品牌") == b))
    for brand in brands:
        subset = [r for r in main if r.get("品牌") == brand]
        mix = Counter(r.get("sentiment_mix") or "" for r in subset)
        row = [brand, len(subset), mix.get("好评", 0), mix.get("差评", 0), mix.get("混合", 0), mix.get("信息不足", 0)]
        for name in neg_l1:
            prefix = f"差评|{name}"
            hit = sum(1 for r in subset if any(r.get(c) and str(c).startswith(prefix) for c in cols))
            row.append(hit)
        ws.append(row)
    autosize(ws, max_width=40)


def write_asin(wb, recs):
    main = main_set(recs)
    by_listing = defaultdict(list)
    for r in main:
        by_listing[r.get("listing_asin") or r.get("ASIN")].append(r)
    ws = wb.create_sheet("07_ASIN切片")
    header_row(ws, ["listing_asin", "品牌", "n", "好评", "差评", "混合", "站点评分数", "覆盖率"])
    items = sorted(by_listing.items(), key=lambda kv: -len(kv[1]))
    for asin, subset in items:
        mix = Counter(r.get("sentiment_mix") or "" for r in subset)
        listed = subset[0].get("站点评分数") if subset else None
        n = len(subset)
        try:
            cov = round(n / float(listed), 4) if listed else None
        except (TypeError, ValueError, ZeroDivisionError):
            cov = None
        ws.append(
            [
                asin,
                subset[0].get("品牌") if subset else "",
                n,
                mix.get("好评", 0),
                mix.get("差评", 0),
                mix.get("混合", 0),
                listed,
                cov,
            ]
        )
    autosize(ws, max_width=24)


def write_source(wb, recs):
    main = main_set(recs)
    ws = wb.create_sheet("08_源表字段")
    header_row(ws, ["维度", "取值", "n"])
    ws.append(["VP", "Y", sum(1 for r in main if r.get("flag_vp") == "Y")])
    ws.append(["Vine", "Y", sum(1 for r in main if r.get("flag_vine") == "Y")])
    ws.append(["有图", "Y", sum(1 for r in main if r.get("flag_has_image") == "Y")])
    ws.append(["有视频", "Y", sum(1 for r in main if r.get("flag_has_video") == "Y")])
    ws.append(["有赞同", ">0", sum(1 for r in main if r.get("赞同数") not in (0, "0", "", None))])
    for k, n in Counter(r.get("所属国家") or "NA" for r in main).most_common():
        ws.append(["国家", k, n])
    for k, n in Counter(str(r.get("评论时间") or "")[:4] or "NA" for r in main).most_common():
        ws.append(["年份", k, n])
    autosize(ws)


def write_tree_derived_detail(wb, header, recs):
    tree = build_tree(header, recs)
    ws = wb.create_sheet("10_详细分析")
    header_row(ws, ["板块", "极性", "一级", "二级", "n"])
    for _sid, no, title, pol, _lead in SECTION_SPEC:
        if pol == "both":
            pols = ("好评", "差评")
        else:
            pols = (pol,)
        for p in pols:
            for l1, l1n in sorted_items(tree[p]["l1"]):
                for l2, l2n in sorted_items(l1n["l2"]):
                    ws.append([f"{no} {title}", p, l1, l2, len(l2n["row_ids"])])
    autosize(ws, max_width=36)


def write_tree_sheet(wb, header, recs):
    tree = build_tree(header, recs)
    ws = wb.create_sheet("13_层级透视")
    cols = ["层级", "极性", "一级", "一级频次", "二级", "二级频次", "三级", "三级频次", "原文摘录", "频次", "品牌", "星级", "row_id"]
    header_row(ws, cols)
    ws.sheet_properties.outlinePr.summaryBelow = False

    def add(level_name, outline, hidden, values):
        ws.append(values)
        r = ws.max_row
        ws.row_dimensions[r].outline_level = outline
        ws.row_dimensions[r].hidden = hidden
        for cell in ws[r]:
            cell.fill = LVL_FILL[level_name]
            cell.font = LVL_FONT[level_name]
            cell.alignment = WRAP
        if level_name == "摘录":
            ws.row_dimensions[r].height = 28

    for pol in ("好评", "差评"):
        pnode = tree[pol]
        pn = len(pnode["row_ids"])
        add("极性", 0, False, ["极性", pol, "", pn, "", "", "", "", "", pn, "", "", ""])
        for l1, l1n in sorted_items(pnode["l1"]):
            n1 = len(l1n["row_ids"])
            add("一级", 1, False, ["一级", pol, l1, n1, "", "", "", "", "", n1, "", "", ""])
            for l2, l2n in sorted_items(l1n["l2"]):
                n2 = len(l2n["row_ids"])
                add("二级", 2, True, ["二级", pol, l1, n1, l2, n2, "", "", "", n2, "", "", ""])
                if l2n["l3"]:
                    for l3, l3n in sorted_items(l2n["l3"]):
                        n3 = len(l3n["row_ids"])
                        add("三级", 3, True, ["三级", pol, l1, n1, l2, n2, l3, n3, "", n3, "", "", ""])
                        for item in l3n["excerpts"]:
                            for ex in item["excerpts"]:
                                add(
                                    "摘录",
                                    4,
                                    True,
                                    ["摘录", pol, l1, n1, l2, n2, l3, n3, ex, 1, item["brand"], item["star"], item["row_id"]],
                                )
                for item in l2n["excerpts"]:
                    for ex in item["excerpts"]:
                        add(
                            "摘录",
                            3 if not l2n["l3"] else 4,
                            True,
                            ["摘录", pol, l1, n1, l2, n2, "", "", ex, 1, item["brand"], item["star"], item["row_id"]],
                        )
    ws.column_dimensions["I"].width = 72
    ws.auto_filter.ref = f"A1:M{ws.max_row}"
    ws.freeze_panes = "A2"


def write_coverage(wb, cov_h, cov_recs):
    if not cov_h:
        return
    ws = wb.create_sheet("12_覆盖率")
    header_row(ws, cov_h)
    for rec in cov_recs:
        ws.append([rec.get(h) for h in cov_h])
    autosize(ws, max_width=48)


def pick_excerpts(items, zh, limit=5):
    texts = []
    seen = set()
    for item in items:
        for ex in item.get("excerpts") or []:
            t = ex.strip()
            if not t or t in seen:
                continue
            seen.add(t)
            texts.append(t)
    scored = sorted(texts, key=lambda t: (abs(len(t) - 120), -len(t)))
    mid = [t for t in scored if 28 <= len(t) <= 280]
    chosen = (mid or scored)[:limit]
    return [{"en": t, "cn": zh.get(t, t)} for t in chosen]


def pack_l1(l1, node, zh):
    l2s = []
    for l2, l2n in sorted_items(node["l2"]):
        children = []
        if l2n["l3"]:
            for l3, l3n in sorted_items(l2n["l3"]):
                children.append({"name": l3, "n": len(l3n["row_ids"]), "excerpts": pick_excerpts(l3n["excerpts"], zh)})
        l2s.append(
            {
                "name": l2,
                "n": len(l2n["row_ids"]),
                "desc": f"{len(l2n['row_ids'])} 条评论提到这一点。",
                "excerpts": pick_excerpts(l2n["excerpts"], zh),
                "l3": children,
            }
        )
    return {"name": l1, "n": len(node["row_ids"]), "l2": l2s}


def collect_pol(tree, pol, zh):
    out = []
    for l1, node in sorted_items(tree[pol]["l1"]):
        packed = pack_l1(l1, node, zh)
        packed["polarity"] = pol
        out.append(packed)
    return out


def build_source_data(recs):
    main = main_set(recs)
    stars, n, rating = star_stats(main)
    mix = Counter(r.get("sentiment_mix") or "" for r in main)
    countries = Counter(r.get("所属国家") or "NA" for r in main)
    months = defaultdict(lambda: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, "t": 0})
    for r in main:
        raw = str(r.get("评论时间") or "")
        month = raw[:7] if len(raw) >= 7 else ""
        try:
            s = int(r.get("星级") or 0)
        except (TypeError, ValueError):
            s = 0
        if month and s in (1, 2, 3, 4, 5):
            months[month][s] += 1
            months[month]["t"] += 1
    time_rows = []
    for m in sorted(months):
        row = {"m": m, "t": months[m]["t"]}
        for s in (1, 2, 3, 4, 5):
            row[str(s)] = months[m][s]
        time_rows.append(row)
    brands = []
    by_brand = defaultdict(list)
    for r in main:
        by_brand[r.get("品牌") or ""].append(r)
    for name, subset in sorted(by_brand.items(), key=lambda kv: -len(kv[1])):
        if not name:
            continue
        mx = Counter(x.get("sentiment_mix") or "" for x in subset)
        brands.append(
            {
                "name": name,
                "n": len(subset),
                "pos": mx.get("好评", 0),
                "neg": mx.get("差评", 0),
                "mix": mx.get("混合", 0),
            }
        )
    asins = []
    by_asin = defaultdict(list)
    for r in main:
        by_asin[r.get("listing_asin") or r.get("ASIN") or ""].append(r)
    for asin, subset in sorted(by_asin.items(), key=lambda kv: -len(kv[1])):
        if not asin:
            continue
        asins.append({"asin": asin, "name": subset[0].get("品牌") or "", "n": len(subset)})
    helpful = sum(1 for r in main if r.get("赞同数") not in (0, "0", "", None))
    return {
        "n": len(main),
        "rating": rating,
        "stars": [{"star": s, "n": stars[s], "pct": round(100 * stars[s] / n)} for s in (5, 4, 3, 2, 1)],
        "helpful": helpful,
        "mix": {
            "好评": mix.get("好评", 0),
            "差评": mix.get("差评", 0),
            "混合": mix.get("混合", 0),
            "信息不足": mix.get("信息不足", 0),
        },
        "vp": sum(1 for r in main if r.get("flag_vp") == "Y"),
        "vine": sum(1 for r in main if r.get("flag_vine") == "Y"),
        "text_only": sum(1 for r in main if r.get("flag_has_image") != "Y" and r.get("flag_has_video") != "Y"),
        "image": sum(1 for r in main if r.get("flag_has_image") == "Y"),
        "video": sum(1 for r in main if r.get("flag_has_video") == "Y"),
        "countries": [[k, v] for k, v in countries.most_common()],
        "brands": brands,
        "asins": asins,
        "time": time_rows,
    }


def default_direction(tree, zh):
    pains = []
    for l1, l1n in sorted_items(tree["差评"]["l1"]):
        for l2, l2n in sorted_items(l1n["l2"]):
            pains.append(
                {
                    "name": l2,
                    "n": len(l2n["row_ids"]),
                    "polarity": "差评",
                    "l1": l1,
                    "improve": [],
                    "product": [],
                    "excerpts": pick_excerpts(l2n["excerpts"], zh, limit=2),
                }
            )
    return {"highlights": collect_pol(tree, "好评", zh), "pains": pains[:12]}


def emit_board_data(out_js: Path, header, recs, meta, zh, direction):
    tree = build_tree(header, recs)
    sections = []
    for sid, no, title, pol, lead in SECTION_SPEC:
        if pol == "both":
            l1 = collect_pol(tree, "好评", zh) + collect_pol(tree, "差评", zh)
        else:
            l1 = collect_pol(tree, pol, zh)
        sections.append({"id": sid, "no": no, "title": title, "lead": lead, "l1": l1})
    if direction is None:
        direction = default_direction(tree, zh)
    source = build_source_data(recs)
    parts = [
        "window.PAGE_META = " + json.dumps(meta, ensure_ascii=False) + ";\n",
        "window.BOARD_DATA = " + json.dumps(sections, ensure_ascii=False) + ";\n",
        "window.DIRECTION_DATA = " + json.dumps(direction, ensure_ascii=False) + ";\n",
        "window.SOURCE_DATA = " + json.dumps(source, ensure_ascii=False) + ";\n",
    ]
    out_js.write_text("".join(parts), encoding="utf-8")


def load_json(path: Path | None):
    if not path:
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def main():
    here = Path(__file__).resolve().parent
    skill_root = here.parent
    parser = argparse.ArgumentParser(description="Build review delivery xlsx and board-data.js")
    parser.add_argument("--master", required=True, help="reviews_master.xlsx")
    parser.add_argument("--out-dir", required=True, help="4 Reports directory")
    parser.add_argument("--html-template", default=str(skill_root / "templates" / "review-analysis.html"))
    parser.add_argument("--category", required=True)
    parser.add_argument("--marketplace", default="US")
    parser.add_argument("--date", default="")
    parser.add_argument("--title", default="Amazon Review Analysis")
    parser.add_argument("--excerpts-cn", default="", help="optional {en: cn} json")
    parser.add_argument("--direction-json", default="", help="optional DIRECTION_DATA json")
    args = parser.parse_args()

    master = Path(args.master).expanduser()
    out_dir = Path(args.out_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    date = args.date or datetime.now().strftime("%Y/%m/%d")
    meta = {
        "title": args.title,
        "category": args.category,
        "marketplace": args.marketplace,
        "date": date,
    }
    zh = load_json(Path(args.excerpts_cn).expanduser()) if args.excerpts_cn else {}
    zh = zh or {}
    direction = load_json(Path(args.direction_json).expanduser()) if args.direction_json else None

    header, recs, tax_h, tax_recs, cov_h, cov_recs = load_master(master)
    wb = openpyxl.Workbook()
    write_readme(wb, meta, recs)
    write_master(wb, header, recs)
    write_taxonomy(wb, tax_h, tax_recs)
    write_l1(wb, header, recs)
    write_l2(wb, tax_recs)
    write_stars(wb, recs)
    write_brand(wb, header, recs)
    write_asin(wb, recs)
    write_source(wb, recs)
    write_tree_derived_detail(wb, header, recs)
    write_coverage(wb, cov_h, cov_recs)
    write_tree_sheet(wb, header, recs)

    slug = args.category.replace(" ", "-")
    stamp = date.replace("/", "")
    xlsx = out_dir / f"{slug}-Review-Analysis-{stamp}.xlsx"
    wb.save(xlsx)

    emit_board_data(out_dir / "board-data.js", header, recs, meta, zh, direction)
    html_src = Path(args.html_template).expanduser()
    if html_src.exists():
        shutil.copy2(html_src, out_dir / "review-analysis.html")
    print(f"wrote {xlsx}")
    print(f"wrote {out_dir / 'board-data.js'}")
    print(f"wrote {out_dir / 'review-analysis.html'}")


if __name__ == "__main__":
    main()
